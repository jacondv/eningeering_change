import base64
import csv
import io
import json

from odoo import _, api, fields, models
from odoo.exceptions import UserError

CONFIG_PARAM_KEY = 'qc_checksheet.panel_import_column_mapping'


class QcChecksheetImportWizard(models.TransientModel):
    """Bulk Excel/CSV import of Panel Lines (techspec §5). Column names are
    user-configurable (not hard-coded) and remembered per-database for the
    next import."""
    _name = 'qc.checksheet.import.wizard'
    _description = 'Import Panel Lines'

    checksheet_id = fields.Many2one(
        'qc.checksheet', required=True,
        domain="[('checksheet_type', '=', 'panel_sticker')]")
    import_file = fields.Binary(string='File', required=True)
    import_filename = fields.Char()
    replace_existing = fields.Boolean(
        string='Replace existing lines',
        help="If checked, all current Panel Lines on the check sheet are removed before import. "
             "Otherwise, imported lines are appended.")
    part_number_column = fields.Char(default='part_number', required=True)
    qty_column = fields.Char(default='qty')
    note_column = fields.Char(default='note')

    @api.model
    def default_get(self, fields_list):
        defaults = super().default_get(fields_list)
        mapping = json.loads(
            self.env['ir.config_parameter'].sudo().get_param(CONFIG_PARAM_KEY, '{}'))
        for key in ('part_number_column', 'qty_column', 'note_column'):
            if key in fields_list and mapping.get(key):
                defaults[key] = mapping[key]
        return defaults

    def _remember_mapping(self):
        self.env['ir.config_parameter'].sudo().set_param(CONFIG_PARAM_KEY, json.dumps({
            'part_number_column': self.part_number_column,
            'qty_column': self.qty_column,
            'note_column': self.note_column,
        }))

    def _read_rows(self):
        """Return a list of dicts keyed by lower-cased header name."""
        content = base64.b64decode(self.import_file)
        filename = (self.import_filename or '').lower()
        if filename.endswith('.csv'):
            text = content.decode('utf-8-sig')
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
        elif filename.endswith('.xlsx'):
            try:
                from openpyxl import load_workbook
            except ImportError:
                raise UserError(_("The openpyxl library is required to import .xlsx files."))
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            rows = [[cell if cell is not None else '' for cell in row]
                    for row in sheet.iter_rows(values_only=True)]
        else:
            raise UserError(_("Unsupported file format. Please upload a .xlsx or .csv file."))

        if not rows:
            return []
        header = [str(cell).strip().lower() for cell in rows[0]]
        return [dict(zip(header, (str(cell).strip() if cell != '' else '' for cell in row)))
                for row in rows[1:] if any(cell != '' for cell in row)]

    def action_import(self):
        self.ensure_one()
        rows = self._read_rows()
        part_col = self.part_number_column.strip().lower()
        qty_col = (self.qty_column or '').strip().lower()
        note_col = (self.note_column or '').strip().lower()

        Product = self.env['product.product']
        PanelLine = self.env['qc.checksheet.panel.line']
        if self.replace_existing:
            self.checksheet_id.panel_line_ids.unlink()

        sequence = 10
        created = self.env['qc.checksheet.panel.line']
        for row in rows:
            code = row.get(part_col, '')
            if not code:
                continue
            product = Product.search([('default_code', '=', code)], limit=1)
            qty = row.get(qty_col, '') if qty_col else ''
            note = row.get(note_col, '') if note_col else ''
            vals = {
                'checksheet_id': self.checksheet_id.id,
                'sequence': sequence,
                'qty': float(qty) if qty else 1.0,
                'note': note,
            }
            if product:
                vals.update({
                    'part_number': product.id,
                    'description': product.name,
                    'image': product.image_1920,
                })
            else:
                vals.update({
                    'description': _("Part Number not found: %s") % code,
                    'note': (note + ' ' if note else '') + _("[Not found in products, please review]"),
                })
            created |= PanelLine.create(vals)
            sequence += 10

        self._remember_mapping()

        if not created:
            raise UserError(_("No rows could be imported. Check the column mapping and file content."))

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'qc.checksheet',
            'res_id': self.checksheet_id.id,
            'view_mode': 'form',
            'target': 'current',
        }
